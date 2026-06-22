import io
import secrets
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..auth import get_current_user
from ..config import get_settings
from ..database import get_db
from ..models import AgentResult, Batch, CallListItem, Candidate, ChatSession, ChatTemplate, Company, User, WPRecord

router = APIRouter(tags=["review"], dependencies=[Depends(get_current_user)])


class CorrectieBody(BaseModel):
    wp_waarde: int
    reden: str | None = None


class BellijstBody(BaseModel):
    reden: str | None = None


class BellijstUpdate(BaseModel):
    status: str | None = None
    notities: str | None = None
    resultaat_wp: int | None = None


def _maak_wp_record(db: Session, cand: Candidate, wp: int, status: str, user: User) -> WPRecord:
    comp = db.get(Company, cand.company_id)
    batch = db.get(Batch, cand.batch_id)
    ar = db.get(AgentResult, cand.gekozen_agent_result) if cand.gekozen_agent_result else None
    rec = WPRecord(company_id=comp.id, candidate_id=cand.id, wp_waarde=wp,
                   wp_jaar=batch.jaar, bron_type=(ar.bron_type if ar else "handmatig"),
                   bron_url=(ar.bron_url if ar else None), status=status,
                   goedgekeurd_door=user.id,
                   goedgekeurd_op=datetime.utcnow())
    db.add(rec)
    return rec


@router.post("/candidates/{candidate_id}/approve")
def approve(candidate_id: str, db: Session = Depends(get_db),
            current_user: User = Depends(get_current_user)):
    cand = db.get(Candidate, candidate_id)
    if not cand:
        raise HTTPException(404, "candidate niet gevonden")
    if cand.wp_kandidaat is None:
        raise HTTPException(422, "geen kandidaat-waarde om goed te keuren")
    cand.status = "approved"
    rec = _maak_wp_record(db, cand, cand.wp_kandidaat, "reviewed", current_user)
    db.commit()
    return {"wp_record_id": rec.id, "wp_waarde": rec.wp_waarde}


@router.post("/candidates/{candidate_id}/correct")
def correct(candidate_id: str, body: CorrectieBody, db: Session = Depends(get_db),
            current_user: User = Depends(get_current_user)):
    cand = db.get(Candidate, candidate_id)
    if not cand:
        raise HTTPException(404, "candidate niet gevonden")
    cand.status = "corrected"
    cand.reconciliatie_reden = (cand.reconciliatie_reden or "") + f" | correctie: {body.reden or 'handmatig'}"
    rec = _maak_wp_record(db, cand, body.wp_waarde, "corrected", current_user)
    db.commit()
    return {"wp_record_id": rec.id, "wp_waarde": rec.wp_waarde}


@router.post("/candidates/{candidate_id}/bellijst")
def zet_op_bellijst(candidate_id: str, body: BellijstBody | None = None,
                    db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    cand = db.get(Candidate, candidate_id)
    if not cand:
        raise HTTPException(404, "candidate niet gevonden")
    comp = db.get(Company, cand.company_id)
    telefoonnummer = comp.enrichment.telefoonnummer if comp.enrichment else None
    reden = (body.reden if body else None) or cand.reconciliatie_reden or "reviewer vraagt belactie"
    cand.status = "to_call"
    item = CallListItem(company_id=comp.id, telefoonnummer=telefoonnummer, reden=reden,
                        toegewezen_aan=current_user.id)
    db.add(item)
    db.commit()
    return {"call_list_id": item.id, "status": cand.status}


@router.post("/candidates/{candidate_id}/create-chat")
async def create_chat_session(candidate_id: str, db: Session = Depends(get_db),
                               current_user: User = Depends(get_current_user)):
    """Maakt een gerichte chat-sessie aan, verstuurt uitnodigingsmail (indien Resend geconfigureerd)."""
    cand = db.get(Candidate, candidate_id)
    if not cand:
        raise HTTPException(404, "candidate niet gevonden")
    default_template = db.query(ChatTemplate).filter_by(is_default=True).first()
    vragen = default_template.vragen if default_template else None

    token = secrets.token_urlsafe(32)
    session = ChatSession(company_id=cand.company_id, token_hash=token,
                          variant="gericht", pre_fill_wp=cand.wp_kandidaat, vragen=vragen)
    db.add(session)
    cand.status = "to_chat"
    db.commit()

    settings = get_settings()
    chat_url = f"{settings.frontend_url}/?chat={token}"
    comp = db.get(Company, cand.company_id)
    email = comp.enrichment.email if comp and comp.enrichment else None

    email_sent = False
    email_recipient = None
    if settings.resend_api_key and (email or settings.email_demo_recipient):
        try:
            from ..email import send_chat_invitation
            await send_chat_invitation(email or "", comp.naam if comp else "", chat_url)
            email_sent = True
            email_recipient = settings.email_demo_recipient or email
        except Exception:
            pass  # email mislukt — frontend valt terug op mailto

    return {"session_id": session.id, "chat_url": chat_url,
            "email": email, "email_sent": email_sent, "email_recipient": email_recipient}


@router.post("/batches/{batch_id}/approve-all-green")
def bulk_approve(batch_id: str, db: Session = Depends(get_db),
                 current_user: User = Depends(get_current_user)):
    """Bulk-goedkeuring van alle 🟢-records (doc §11)."""
    n = 0
    for cand in db.query(Candidate).filter_by(batch_id=batch_id, confidence_label="hoog",
                                              status="pending"):
        if cand.wp_kandidaat is not None:
            cand.status = "approved"
            _maak_wp_record(db, cand, cand.wp_kandidaat, "reviewed", current_user)
            n += 1
    db.commit()
    return {"goedgekeurd": n}


@router.get("/batches/{batch_id}/bellijst")
def get_bellijst(batch_id: str, db: Session = Depends(get_db)):
    """Alle bellijst-items voor een batch, met bedrijfsinfo."""
    items = (db.query(CallListItem)
             .join(Company, CallListItem.company_id == Company.id)
             .filter(Company.batch_id == batch_id)
             .order_by(CallListItem.created_at)
             .all())
    result = []
    for item in items:
        comp = db.get(Company, item.company_id)
        result.append({
            "id": item.id,
            "company_id": item.company_id,
            "naam": comp.naam,
            "gemeente": comp.gemeente,
            "telefoonnummer": item.telefoonnummer or (comp.enrichment.telefoonnummer if comp.enrichment else None),
            "email": comp.enrichment.email if comp.enrichment else None,
            "reden": item.reden,
            "status": item.status,
            "notities": item.notities,
            "resultaat_wp": item.resultaat_wp,
            "created_at": item.created_at.isoformat() + "Z" if item.created_at else None,
        })
    return result


@router.patch("/bellijst/{item_id}")
def update_bellijst_item(item_id: str, body: BellijstUpdate, db: Session = Depends(get_db)):
    """Status, notities en resultaat-WP bijwerken voor een bellijst-item."""
    item = db.get(CallListItem, item_id)
    if not item:
        raise HTTPException(404, "bellijst-item niet gevonden")
    if body.status is not None:
        item.status = body.status
    if body.notities is not None:
        item.notities = body.notities
    if body.resultaat_wp is not None:
        item.resultaat_wp = body.resultaat_wp
    db.commit()
    return {"id": item.id, "status": item.status}


@router.post("/bellijst/{item_id}/doorvoeren")
def doorvoeren_bellijst(item_id: str, db: Session = Depends(get_db),
                        current_user: User = Depends(get_current_user)):
    """Resultaat-WP van een afgerond bellijst-item doorvoeren als gecorrigeerde waarde."""
    item = db.get(CallListItem, item_id)
    if not item:
        raise HTTPException(404, "bellijst-item niet gevonden")
    if not item.resultaat_wp:
        raise HTTPException(422, "geen resultaat-WP ingevuld om door te voeren")
    comp = db.get(Company, item.company_id)
    cand = comp.candidate if comp else None
    if not cand:
        raise HTTPException(404, "geen kandidaat gevonden voor dit bedrijf")
    cand.status = "corrected"
    cand.reconciliatie_reden = (cand.reconciliatie_reden or "") + \
        f" | belactie {current_user.naam}: {item.resultaat_wp} WP"
    rec = _maak_wp_record(db, cand, item.resultaat_wp, "corrected", current_user)
    db.commit()
    return {"wp_record_id": rec.id, "wp_waarde": rec.wp_waarde}


_ETIL_TEAL = "115E59"
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_BODY_FONT = Font(name="Calibri", size=11)
_STRIPE = PatternFill("solid", fgColor="F0FAFA")
_THIN = Side(style="thin", color="D1D5DB")
_BORDER = Border(bottom=_THIN)


def _style_sheet(ws, headers: list[str], col_widths: list[int]) -> None:
    fill = PatternFill("solid", fgColor=_ETIL_TEAL)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _style_row(ws, row_idx: int, n_cols: int) -> None:
    stripe = row_idx % 2 == 0
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = _BODY_FONT
        cell.alignment = Alignment(vertical="center")
        if stripe:
            cell.fill = _STRIPE
        cell.border = _BORDER


def _wb_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return StreamingResponse(iter([buf.getvalue()]), media_type=media,
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


_CONF_LABELS = {"hoog": "Groen", "middel": "Geel", "laag": "Rood"}


@router.get("/batches/{batch_id}/export.xlsx")
def export_xlsx(batch_id: str, db: Session = Depends(get_db)):
    """Export van alle WP-records voor het Vestigingsregister als Excel."""
    batch = db.get(Batch, batch_id)
    naam = (batch.naam or batch_id) if batch else batch_id

    headers = [
        "Vestigingsnummer", "Naam", "Gemeente", "Adres", "SBI-code", "CB-er", "KvK",
        "WP", "Jaar", "Bron", "Bron-URL", "Status", "Betrouwbaarheid",
        "Man", "Vrouw", "Voltijd", "Deeltijd",
        "Eigen personeel", "Uitzend", "Detachering", "WSW",
        "% op locatie",
    ]
    widths = [18, 34, 18, 28, 10, 10, 14, 7, 6, 14, 40, 14, 14,
              7, 7, 8, 8, 16, 9, 13, 7, 12]

    wb = Workbook()
    ws = wb.active
    ws.title = "Vestigingsregister"
    ws.sheet_properties.tabColor = _ETIL_TEAL
    _style_sheet(ws, headers, widths)

    rows = (db.query(WPRecord, Company, Candidate)
            .join(Company, WPRecord.company_id == Company.id)
            .outerjoin(Candidate, Candidate.company_id == Company.id)
            .filter(Company.batch_id == batch_id)
            .order_by(Company.naam)
            .all())

    for row_idx, (rec, comp, cand) in enumerate(rows, 2):
        pct = f"{round(rec.pct_op_locatie * 100)}%" if rec.pct_op_locatie is not None else None
        values = [
            comp.vestigingsnummer, comp.naam, comp.gemeente, comp.adres,
            comp.sbi_code, comp.cb_er, comp.kvk_nummer,
            rec.wp_waarde, rec.wp_jaar, rec.bron_type, rec.bron_url, rec.status,
            _CONF_LABELS.get(cand.confidence_label, "") if cand else "",
            rec.man, rec.vrouw, rec.voltijd, rec.deeltijd,
            rec.eigen_personeel, rec.uitzend, rec.detachering, rec.wsw,
            pct,
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        _style_row(ws, row_idx, len(headers))

    return _wb_response(wb, f"vestigingsregister_{naam}.xlsx")


@router.get("/batches/{batch_id}/bellijst.xlsx")
def export_bellijst_xlsx(batch_id: str, db: Session = Depends(get_db)):
    """Export van de bellijst als Excel."""
    batch = db.get(Batch, batch_id)
    naam = (batch.naam or batch_id) if batch else batch_id

    headers = ["Naam", "Gemeente", "Telefoonnummer", "WP-kandidaat", "Reden", "Status", "Resultaat WP"]
    widths = [34, 18, 18, 14, 50, 14, 12]

    wb = Workbook()
    ws = wb.active
    ws.title = "Bellijst"
    ws.sheet_properties.tabColor = _ETIL_TEAL
    _style_sheet(ws, headers, widths)

    rows = (db.query(CallListItem, Company, Candidate)
            .join(Company, CallListItem.company_id == Company.id)
            .outerjoin(Candidate, Candidate.company_id == Company.id)
            .filter(Company.batch_id == batch_id)
            .order_by(Company.naam)
            .all())

    for row_idx, (item, comp, cand) in enumerate(rows, 2):
        values = [
            comp.naam, comp.gemeente, item.telefoonnummer,
            cand.wp_kandidaat if cand else None,
            item.reden, item.status, item.resultaat_wp,
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        _style_row(ws, row_idx, len(headers))

    return _wb_response(wb, f"bellijst_{naam}.xlsx")
